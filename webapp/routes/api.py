import asyncio
import os
import zipfile
from copy import copy
from io import BytesIO
from urllib.parse import quote
from uuid import uuid4

import aiofiles.os
import cv2
from aiohttp import web
from openpyxl.drawing.image import Image
from openpyxl.workbook import Workbook

from platerecognizer.api import PlateRecognizerAPI
from utils import write_frame
from webapp.constants import VIDEOFILES_PATH
from webapp.video import Video, flatten

api_routes = web.RouteTableDef()

plate_recognizer_api = PlateRecognizerAPI('06e820d5157f96a75f900232e5ffffc2752269ae', ['ru'])


async def makedirs(path: str) -> str:
    await aiofiles.os.makedirs(path, exist_ok=True)
    return path


@api_routes.post('/detect')
async def detect(request: web.Request):
    data = await request.json()
    result = plate_recognizer_api.detect(data['image_filepath'])
    return web.json_response(result)


@api_routes.post('/find_unique_plates')
async def find_unique_plates(request: web.Request):
    filename = (await request.json())['filename']
    recognizing = copy(request.app.recognizing_video)
    added_to_queue = request.app.add_to_queue(filename)
    request.app.save_queue()
    return web.json_response({'added_to_queue': added_to_queue, 'recognizing': recognizing})


@api_routes.post('/upload_video')
async def upload_video(request: web.Request):
    reader = await request.multipart()
    field = await reader.next()
    os.makedirs(VIDEOFILES_PATH, exist_ok=True)
    filepath = VIDEOFILES_PATH + '/' + field.filename
    async with aiofiles.open(filepath, 'wb') as f:
        while True:
            chunk = await field.read_chunk()
            if not chunk:
                break
            await f.write(chunk)

    cv2_video = cv2.VideoCapture(filepath)
    ret, frame = cv2_video.read()
    total_frames = int(cv2_video.get(cv2.CAP_PROP_FRAME_COUNT))
    fps = int(cv2_video.get(cv2.CAP_PROP_FRAME_COUNT))
    cv2_video.release()
    new_height = 100
    ratio = new_height / frame.shape[0]
    new_width = int(frame.shape[1] * ratio)
    preview = cv2.resize(frame, (new_width, new_height))
    write_frame(f'preview_{field.filename}.jpg', preview)

    video = Video(field.filename, total_frames, fps)
    await video.save()
    request.app['videos'].append(video)

    return web.json_response(video.tojson())


@api_routes.post('/get_video')
async def get_video(request: web.Request):
    filename = (await request.json())['filename']
    video = request.app.get_video(filename)
    return web.json_response(video.tojson())


@api_routes.post('/get_queue')
async def get_queue(request: web.Request):
    return web.json_response(request.app.queue)


@api_routes.post('/get_videos')
async def get_videos(request: web.Request):
    json_response = [flatten(v) for v in request.app['videos']]
    return web.json_response(json_response)


@api_routes.post('/download_excel')
async def download_excel(request: web.Request):
    print('enter')
    filename = (await request.json())['filename']
    video = request.app.get_video(filename)

    wb = Workbook()
    ws = wb.active

    # Добавление заголовков столбцов
    headers = ['Номер', 'Изображение']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header

    # Добавление данных
    for row_num, plate in enumerate(video.filtered_plates, 2):
        ws.cell(row=row_num, column=1, value=plate.censored_text)
        img_data = open('html/img/' + plate.filename, 'rb').read()
        img = Image(BytesIO(img_data))
        img.fitToCell = True
        img.height = 50
        img.width = 200
        ws.column_dimensions['B'].width = 200
        ws.row_dimensions[row_num].height = 50
        ws.add_image(img, f'B{row_num}')

    excel_data = BytesIO()
    wb.save(excel_data)
    excel_data.seek(0)

    # Set the filename in the Content-Disposition header
    encoded_filename = quote(filename)
    content_disposition = f'attachment; filename="{encoded_filename}.xlsx"'
    print(content_disposition)

    # Return the workbook as a response with the appropriate headers
    return web.Response(
        body=excel_data,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            'Content-Disposition': content_disposition
        })


@api_routes.post('/download_txt')
async def download_txt(request: web.Request):
    filename = (await request.json())['filename']
    video = request.app.get_video(filename)

    plates_text = []
    predictions = {
        '79_': [0, 7, 9],
        '19_': [7, 9],
        '17_': [7],
        '7_': [7],
        '9_': [7, 9],
        '27_': [7, 9],
        '77_': [7],
        '97_': [7]
    }
    for plate in video.filtered_plates:
        if plate.censored_text.endswith('79_'):
            for plate_ending, predicted_ending_numbers in predictions.items():
                if plate.censored_text.endswith(plate_ending):
                    for predicted_number in predicted_ending_numbers:
                        plates_text.append(plate.censored_text.replace('_', str(predicted_number)))
    print(plates_text)

    # Set the filename in the Content-Disposition header
    encoded_filename = quote(filename)
    content_disposition = f'attachment; filename="{encoded_filename}.txt"'
    print(content_disposition)

    # Return the workbook as a response with the appropriate headers
    return web.Response(
        body=' '.join(plates_text),
        content_type='text/plain',
        headers={
            'Content-Disposition': content_disposition
        })


@api_routes.post('/delete_video')
async def delete_video(request: web.Request):
    filename = (await request.json())['filename']
    success = await request.app.delete_video(filename)
    return web.json_response({'success': success})


@api_routes.post('/download_cars')
async def download_cars(request: web.Request):
    # Получить имя файла из тела запроса
    data = await request.json()
    filename = data['filename']
    video = request.app.get_video(filename)

    # Создать временный файл для хранения архива
    zip_filename = f'{str(uuid4())}.zip'
    zip_filepath = f'.temp/{zip_filename}'
    os.makedirs(os.path.dirname(zip_filepath), exist_ok=True)

    # Создать зип архив и добавить файлы в него
    with zipfile.ZipFile(zip_filepath, 'w') as zip_file:
        files = [f for f in os.listdir('html/img') if filename in f and f.startswith('car')]
        for plate in video.filtered_plates:
            for file in files:
                if plate.censored_text in file:
                    filepath = f'html/img/{file}'
                    zip_file.write(filepath, os.path.basename(filepath))


    # Открыть архив и считать его содержимое
    with open(zip_filepath, 'rb') as zip_file:
        zip_content = zip_file.read()

    # Удалить временный файл архива
    os.remove(zip_filepath)

    # Отправить архив как ответ
    headers = {
        'Content-Disposition': f'attachment; filename={filename}.zip'
    }
    return web.Response(body=zip_content, content_type='application/zip', headers=headers)


@api_routes.post('/stop_video')
async def stop_video(request: web.Request):
    filename = (await request.json())['filename']
    request.app.stop_recognizing = True
    while request.app.stop_recognizing:
        await asyncio.sleep(1)
    return web.json_response(request.app.get_video(filename).tojson())


@api_routes.post('/reset_video')
async def reset_video(request: web.Request):
    filename = (await request.json())['filename']
    video = request.app.get_video(filename)
    video.reset()
    video = request.app.get_video(filename)
    return web.json_response(video.tojson())
