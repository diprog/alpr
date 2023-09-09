import asyncio
import json
from io import BytesIO

from openpyxl import Workbook
from openpyxl.drawing.image import Image

from webapp import restore
from webapp.video import Video


async def main():
    with open('.data/videos_json/IMG_3139.json', 'r', encoding='utf-8') as f:
        video: Video = restore(json.load(f))
    wb = Workbook()
    ws = wb.active

    # Добавление заголовков столбцов
    headers = ['Номер', 'Изображение']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header

    # Добавление данных
    for row_num, plate in enumerate(video.filtered_plates, 2):
        ws.cell(row=row_num, column=1, value=plate.censored_text)
        img_data = open('html/img/' + plate.filename, 'rb').read()
        img = Image(BytesIO(img_data))
        img.fitToCell = True
        ws.column_dimensions['B'].width = 200
        ws.row_dimensions[row_num].height = 50
        ws.add_image(img, f'B{row_num}')

    wb.save("output.xlsx")


if __name__ == '__main__':
    asyncio.run(main())
